"""Export utilities · convert a pandas DataFrame to XLSX or PDF bytes."""

import io
from typing import Optional

import pandas as pd


def to_xlsx(df: pd.DataFrame, title: str = "Export") -> bytes:
    """Return an Excel (.xlsx) file as bytes.

    The first row is formatted as a bold header with a light-blue fill.

    Args:
        df: DataFrame to export
        title: used as the sheet name (truncated to 31 chars for Excel)

    Returns:
        Raw XLSX bytes
    """
    buffer = io.BytesIO()
    sheet_name = title[:31] if title else "Sheet1"

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)

        # Style the header row
        try:
            from openpyxl.styles import Font, PatternFill, Alignment

            ws = writer.sheets[sheet_name]
            header_fill = PatternFill(
                start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"
            )
            header_font = Font(bold=True)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Auto-fit column widths (approximate)
            for col_cells in ws.columns:
                max_len = max(
                    (len(str(c.value)) if c.value is not None else 0) for c in col_cells
                )
                ws.column_dimensions[col_cells[0].column_letter].width = min(
                    max_len + 4, 50
                )
        except ImportError:
            pass  # openpyxl not available with styles · export without formatting

    return buffer.getvalue()


def to_pdf(df: pd.DataFrame, title: str = "Export") -> bytes:
    """Return a PDF file as bytes with a simple table layout.

    Args:
        df: DataFrame to export
        title: document title printed at the top

    Returns:
        Raw PDF bytes
    """
    try:
        from fpdf import FPDF
    except ImportError as exc:
        raise ImportError(
            "fpdf2 is required for PDF export. Install it with: pip install fpdf2"
        ) from exc

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()

    # Title
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, title, ln=True, align="C")
    pdf.ln(4)

    cols = list(df.columns)
    n_cols = len(cols)
    page_width = pdf.w - 2 * pdf.l_margin
    col_width = page_width / n_cols if n_cols else page_width

    # Header row
    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(189, 215, 238)  # light blue
    for col in cols:
        pdf.cell(col_width, 7, str(col)[:30], border=1, align="C", fill=True)
    pdf.ln()

    # Data rows
    pdf.set_font("Helvetica", "", 7)
    pdf.set_fill_color(255, 255, 255)

    for i, (_, row) in enumerate(df.iterrows()):
        if i >= 500:  # cap at 500 rows in PDF to keep file manageable
            pdf.set_font("Helvetica", "I", 7)
            pdf.cell(
                0, 6,
                f"... {len(df) - 500} additional rows not shown",
                ln=True, align="L",
            )
            break

        fill = i % 2 == 1
        if fill:
            pdf.set_fill_color(242, 242, 242)
        else:
            pdf.set_fill_color(255, 255, 255)

        for col in cols:
            value = str(row[col]) if row[col] is not None else ""
            pdf.cell(col_width, 6, value[:30], border=1, align="L", fill=True)
        pdf.ln()

    return bytes(pdf.output())
